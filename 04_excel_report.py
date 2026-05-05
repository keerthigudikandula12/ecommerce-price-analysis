


import pandas as pd
import pickle
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os

with open("outputs/sql_results.pkl", "rb") as f:
    res = pickle.load(f)
df_clean = pd.read_csv("outputs/clean_data.csv")

wb = Workbook()

# ── colour palette ────────────────────────────────────────
BLUE_DARK  = "1B4F72"
BLUE_MID   = "2E86C1"
BLUE_LIGHT = "EBF5FB"
GREEN      = "3B6D11"
GREEN_LIGHT= "EAF3DE"
GRAY_HEAD  = "F2F3F4"
WHITE      = "FFFFFF"

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_table(ws, df, start_row=1, header_fill_hex=BLUE_MID):
    hf = fill(header_fill_hex)
     # ── Step A: Write the header row ─────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font      = hdr_font()
        cell.fill      = hf
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_border()

    # ── Step B: Write every data row ─────────────

    for row_idx, row in enumerate(df.itertuples(index=False), start_row+1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = cell_font()
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if row_idx % 2 == 0:
                cell.fill = fill(BLUE_LIGHT)
                
    # ── Step C: Set column widths ─────────────────

    for col_idx in range(1, len(df.columns)+1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

# ── SHEET 1: Summary KPIs ─────────────────────────────────
ws1 = wb.active
ws1.title = "Summary KPIs"

ws1["A1"] = "E-Commerce Price Tracker — Summary"
ws1["A1"].font = Font(bold=True, size=16, color=BLUE_DARK, name="Calibri")
ws1.merge_cells("A1:F1")
ws1["A1"].alignment = Alignment(horizontal="left")

kpis = [
    ("Total Products Analysed", len(df_clean)),
    ("Avg Discount %",          f"{df_clean['discount_pct'].mean():.1f}%"),
    ("Avg Discounted Price ₹",  f"₹{df_clean['discounted_price'].mean():,.0f}"),
    ("Max Savings on 1 Product",f"₹{df_clean['savings'].max():,.0f}"),
    ("Products with 50%+ Off",  int((df_clean['discount_pct'] >= 50).sum())),
    ("Avg Rating",              f"{df_clean['rating'].mean():.2f} / 5.0"),
]

ws1["A3"] = "Key Metrics"
ws1["A3"].font = Font(bold=True, size=11, color=WHITE, name="Calibri")
ws1["A3"].fill = fill(BLUE_MID)
ws1.merge_cells("A3:F3")

for i, (label, value) in enumerate(kpis, 4):
    ws1.cell(row=i, column=1, value=label).font = cell_font(bold=True)
    val_cell = ws1.cell(row=i, column=2, value=value)
    val_cell.font = Font(bold=True, size=12, color=BLUE_MID, name="Calibri")
    if i % 2 == 0:
        ws1.cell(row=i, column=1).fill = fill(GRAY_HEAD)
        val_cell.fill = fill(GRAY_HEAD)
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22

# ── SHEET 2: Discount by Category ────────────────────────
ws2 = wb.create_sheet("Discount by Category")
ws2["A1"] = "Average Discount % by Category"
ws2["A1"].font = Font(bold=True, size=13, color=BLUE_DARK, name="Calibri")
write_table(ws2, res["avg_discount_by_cat"], start_row=3)

# add colour scale on discount column
ws2.conditional_formatting.add(
    f"B4:B{3+len(res['avg_discount_by_cat'])}",
    ColorScaleRule(
        start_type="min", start_color="EAF3DE",
        end_type="max",   end_color="3B6D11"
    )
)

# ── SHEET 3: Top Savings ──────────────────────────────────
ws3 = wb.create_sheet("Top Savings")
ws3["A1"] = "Top 10 Products by Absolute Savings (₹)"
ws3["A1"].font = Font(bold=True, size=13, color=BLUE_DARK, name="Calibri")
write_table(ws3, res["top_savings"], start_row=3, header_fill_hex=GREEN)

# ── SHEET 4: Best Value ───────────────────────────────────
ws4 = wb.create_sheet("Best Value Products")
ws4["A1"] = "Best Value — High Rating + High Discount"
ws4["A1"].font = Font(bold=True, size=13, color=BLUE_DARK, name="Calibri")
write_table(ws4, res["best_value"], start_row=3, header_fill_hex=GREEN)

# ── SHEET 5: Value Traps ──────────────────────────────────
ws5 = wb.create_sheet("Value Traps")
ws5["A1"] = "Value Traps — High Discount but Low Rating"
ws5["A1"].font = Font(bold=True, size=13, color="A32D2D", name="Calibri")
write_table(ws5, res["value_traps"], start_row=3,
            header_fill_hex="A32D2D")

# ── SHEET 6: Raw Clean Data ───────────────────────────────
ws6 = wb.create_sheet("Clean Data")
cols_to_export = [
    "product_short","main_category","discounted_price",
    "actual_price","discount_pct","savings","rating","rating_count"
]
write_table(ws6, df_clean[cols_to_export].head(200), start_row=1)

# ── Save ──────────────────────────────────────────────────
path = "outputs/dashboard.xlsx"
wb.save(path)
print(f"Excel dashboard saved: {path}")
print(f"Sheets: Summary KPIs | Discount by Category | "
      f"Top Savings | Best Value | Value Traps | Clean Data")

